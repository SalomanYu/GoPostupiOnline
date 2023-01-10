from sql_storage import getPrograms, getSpecs
from typing import NamedTuple
import openpyxl

class Specialization(NamedTuple):
	Id: str
	Name: str
	VuzId: str

def specializationHasProfessions(specId: str) -> bool:
	try:
		next(program[ProgramHasProfessionColumn] for program in getPrograms() if program[ProgramSpecColumn] == specId and program[ProgramHasProfessionColumn] == 1)
		return True
	except:
		return False

SpecIdColumn = 0
SpecVuzColumn = 1
SpecNameColumn = 2

ProgramIdColumn = 0
ProgramSpecColumn = 1
ProgramHasProfessionColumn = -2


specs = getSpecs()
specsWithoutProfessions: list[Specialization] = [Specialization("1232", "Fdsfsd", "sdsadsa"), Specialization("1232", "Fdsfsd", "sdsadsa"), Specialization("1232", "Fdsfsd", "sdsadsa")]
length = len(specs)
book = openpyxl.Workbook()
sheet = book.active


row = 1
sheet.cell(row, 1).value = "id"
sheet.cell(row, 2).value = "name"
sheet.cell(row, 3).value = "vuz"
for index, item in enumerate(specs):
	print("Осталось:", length - index)
	if not specializationHasProfessions(item[SpecIdColumn]):
		row += 1
		sheet.cell(row, 1).value = item[SpecIdColumn]
		sheet.cell(row, 2).value = item[SpecVuzColumn]
		sheet.cell(row, 3).value = item[SpecNameColumn]

# for item in specsWithoutProfessions:
book.save("specializationWithoutProfessions.xlsx")